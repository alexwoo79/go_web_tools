#!/usr/bin/env python3
"""Excel 结构解析共享库（analyze_excel.py / excel_to_yaml.py 共用）。

能力：加载 xlsx/csv、识别表头（支持合并标题/多级表头）、解析合并单元格分组标签、
按关键字与样例值推断字段类型、中文转拼音字段名。
"""

from __future__ import annotations

import csv
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None

try:
    from pypinyin import lazy_pinyin
except ImportError:  # pragma: no cover
    lazy_pinyin = None


MAX_SELECT_OPTIONS = 12
MAX_OPTION_LENGTH = 20
MIN_SELECT_DISTINCT = 2
FOOTER_KEYWORDS = ("签字", "签名", "合计", "总分", "总得分", "总评", "考核结果", "总计")


def cell_text(v) -> str:
    """单元格值转去空格文本。"""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def load_rows(path: str, sheet_name: str | None = None):
    """加载文件，返回 (sheets, active_sheet, rows, merged_map)。

    rows: list[list[str]]，每行长度不固定，越界访问需自行判断。
    merged_map: {(row, col): top_left_value}，1 起，合并范围内除左上角外都能取到分组标题。
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"文件不存在: {path}")

    if path.suffix.lower() == ".csv":
        with open(path, encoding="utf-8-sig", newline="") as f:
            rows = [list(r) for r in csv.reader(f)]
        return [path.name], path.name, rows, {}

    if path.suffix.lower() == ".xls":
        raise ValueError("暂不支持旧版 .xls，请另存为 .xlsx 后重试")

    if openpyxl is None:
        raise RuntimeError("缺少 openpyxl，请先执行: pip install -r requirements.txt")

    wb = openpyxl.load_workbook(path, data_only=True)
    sheets = wb.sheetnames
    if not sheets:
        raise ValueError("Excel 中没有工作表")

    if sheet_name:
        if sheet_name not in sheets:
            raise ValueError(f"工作表 {sheet_name} 不存在，可选: {', '.join(sheets)}")
        ws = wb[sheet_name]
    else:
        ws = wb[sheets[0]]

    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([cell_text(c) for c in row])

    merged = {}
    for rng in ws.merged_cells.ranges:
        val = cell_text(ws.cell(rng.min_row, rng.min_col).value)
        if not val:
            continue
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                merged.setdefault((r, c), val)

    return sheets, ws.title, rows, merged


def row_stats(row) -> tuple[int, list[str], str]:
    cells = [cell_text(c) for c in row]
    non_empty = [c for c in cells if c]
    return len(non_empty), cells, " ".join(non_empty)


def detect_header(rows: list[list[str]]):
    """识别表头，返回 dict(title, groups, header_row, header_idx)。

    groups: [{row, idx}]，字段行之上的分组行（自下而上使用）。
    规则：第一个只有 1 个非空单元格的行是大标题；第一个 >=2 非空单元格的行是
    表头块起点；若下一行非空单元格更多，则当前行是分组行、下一行是字段行。
    """
    candidates = []
    for i, row in enumerate(rows):
        cnt, cells, text = row_stats(row)
        if cnt == 0:
            continue
        candidates.append({"idx": i, "row": cells, "cnt": cnt, "text": text})

    if not candidates:
        return None

    # 跳过「标签：值」信息行；第一个单格行是大标题，后续单格行是合并的分节/分组标签
    title = ""
    leading_groups = []
    block_start = -1
    for i, c in enumerate(candidates):
        if _is_info_row(c["row"]):
            continue
        if c["cnt"] == 1:
            if not title:
                title = c["text"]
            else:
                leading_groups.append({"row": c["row"], "idx": c["idx"]})
            continue
        block_start = i
        break
    if block_start == -1:
        return {"title": title, "groups": [], "header_row": [], "header_idx": -1}

    block = [candidates[block_start]]
    if (
        block_start + 1 < len(candidates)
        and block[0]["cnt"] <= 2
        and candidates[block_start + 1]["cnt"] >= 4
        and candidates[block_start + 1]["cnt"] > block[0]["cnt"]
        and not _is_info_row(candidates[block_start + 1]["row"])
    ):
        block.append(candidates[block_start + 1])

    header = block[-1]
    groups = list(leading_groups)
    if len(block) == 2:
        groups.append({"row": block[0]["row"], "idx": block[0]["idx"]})
    return {
        "title": title,
        "groups": groups,
        "header_row": header["row"],
        "header_idx": header["idx"],
    }


def _is_info_row(row: list[str]) -> bool:
    cells = [cell_text(v) for v in row if cell_text(v)]
    for j, v in enumerate(cells):
        if LABEL_VALUE_RE.match(v) and j + 1 < len(cells):
            return True
    return False


def resolve_label(header, merged: dict, col: int):
    """解析某列表头文字，返回 (label, source)。

    source: header=字段行原文，merged=合并分组标题回退，group=分组行回退，""=无。
    顺序：字段行 → 合并单元格分组标题 → 分组行（自下而上）。
    """
    if col < len(header["header_row"]):
        v = header["header_row"][col]
        if v:
            return v, "header"

    # 字段行为空：查该单元格是否落在合并分组内
    if merged.get((header["header_idx"] + 1, col + 1)):
        return merged[(header["header_idx"] + 1, col + 1)], "merged"

    # 分组行（自下而上）
    for g in reversed(header["groups"]):
        if merged.get((g["idx"] + 1, col + 1)):
            return merged[(g["idx"] + 1, col + 1)], "group"
        if col < len(g["row"]) and g["row"][col]:
            return g["row"][col], "group"
    return "", ""


def detect_info_fields(rows: list[list[str]], header_idx: int):
    """识别表头上方的「标签：值」信息行（如 部门：设计研究部），返回字段列表。

    返回 [{label, name, value}]；value 是模板样例值，仅用于展示，不写入表单默认值。
    """
    out = []
    seen = set()
    for i in range(header_idx):
        cells = [(c, cell_text(v)) for c, v in enumerate(rows[i]) if cell_text(v)]
        for j, (_, v) in enumerate(cells):
            if not LABEL_VALUE_RE.match(v):
                continue
            label = v.rstrip("：: ").strip()
            if not label or label in seen:
                continue
            value = cells[j + 1][1] if j + 1 < len(cells) else ""
            if not value:
                continue
            seen.add(label)
            out.append({"label": label, "name": slugify(label), "value": value})
    return out


REQUIRED_RE = re.compile(r"[*＊]|必填|（必填）|\(必填\)|required", re.IGNORECASE)
NUMBER_RE = re.compile(r"^-?\d+(\.\d+)?$")
DATE_RE = re.compile(r"^\d{4}[-/.]\d{1,2}[-/.]\d{1,2}")
PERCENT_RE = re.compile(r"^-?\d+(\.\d+)?%$")
LABEL_VALUE_RE = re.compile(r"^[^：:\n]{1,10}[：:]$")


def is_numeric(s: str) -> bool:
    return bool(NUMBER_RE.match(s.strip()))


def is_percent(s: str) -> bool:
    return bool(PERCENT_RE.match(s.strip()))


def percent_like(samples: list[str]) -> bool:
    if not samples:
        return False
    return all(is_percent(s) for s in samples)


def infer_options(samples: list[str]) -> list[str] | None:
    """样例值去重后提取有限选项；不满足条件返回 None。"""
    distinct = []
    seen = set()
    for s in samples:
        s = re.sub(r"\s+", "", s)  # 合并单元格文本可能含换行，归一化
        if not s or len(s) > MAX_OPTION_LENGTH or is_numeric(s) or DATE_RE.match(s):
            return None
        if s in seen:
            continue
        seen.add(s)
        distinct.append(s)
    if len(distinct) < MIN_SELECT_DISTINCT or len(distinct) > MAX_SELECT_OPTIONS:
        return None
    return distinct


def infer_type(label: str, samples: list[str]):
    """返回 (type, options, min, max)。"""
    l = label.lower()
    if "日期" in l:
        return "date", None, None, None
    if "时间" in l:
        return "time", None, None, None
    if "邮箱" in l or "邮件" in l or "e-mail" in l:
        return "email", None, None, None
    if "电话" in l or "手机" in l or "联系方式" in l:
        return "tel", None, None, None
    if any(k in l for k in ("得分", "分数", "评分", "分值")):
        if percent_like(samples):
            opts = infer_options(samples)
            return ("select", opts, None, None) if opts else ("text", None, None, None)
        return "number", None, 0, 100
    if any(k in l for k in ("序号", "编号", "工号", "行号")):
        return "number", None, 0, None
    if any(k in l for k in ("权重", "百分比", "比例", "占比")):
        if percent_like(samples):
            opts = infer_options(samples)
            return ("select", opts, None, None) if opts else ("text", None, None, None)
        return "number", None, 0, None
    if any(k in l for k in ("数量", "人数", "年龄", "金额", "合同额", "产值", "额度", "价格", "工作量", "公里", "里程", "时长", "分钟")):
        if percent_like(samples):
            opts = infer_options(samples)
            return ("select", opts, None, None) if opts else ("text", None, None, None)
        return "number", None, 0, None
    if any(k in l for k in ("评价", "意见", "描述", "说明", "内容", "总结", "自评", "建议", "心得", "体会", "汇报", "备注", "评语", "完成情况", "目标", "要点")):
        return "textarea", None, None, None
    if "多选" in l or "勾选" in l:
        return "checkbox", None, None, None
    if any(k in l for k in ("选择", "性别", "部门", "类别", "类型", "状态", "地区", "城市", "学历", "岗位", "职位", "等级", "级别", "是否", "渠道", "方向", "选项")):
        opts = infer_options(samples)
        return ("select", opts, None, None) if opts else ("select", None, None, None)
    if any(k in l for k in ("姓名", "名称", "项目", "指标", "单位", "编号", "单号", "地址", "签字", "签名", "工号")):
        return "text", None, None, None
    opts = infer_options(samples)
    return ("select", opts, None, None) if opts else ("text", None, None, None)


def slugify(label: str) -> str:
    """中文转拼音小写下划线；无 pypinyin 时退化为英文数字，纯中文则返回空。"""
    if lazy_pinyin is None:
        s = re.sub(r"[^a-zA-Z0-9]+", "_", label.lower()).strip("_")
        return s
    out = []
    for ch in label:
        if ch.isascii() and (ch.isalnum()):
            out.append(ch.lower())
        elif "\u4e00" <= ch <= "\u9fff":
            py = lazy_pinyin(ch)[0]
            if py:
                if out and out[-1] != "_":
                    out.append("_")
                out.append(py)
        elif ch in " _-/（）()":
            if out and out[-1] != "_":
                out.append("_")
    return re.sub(r"_+", "_", "".join(out)).strip("_")


def sample_values(
    rows: list[list[str]],
    start: int,
    col: int,
    limit: int = 30,
    header_row: list[str] | None = None,
) -> list[str]:
    """收集某列样例值；跳过空行、页脚行、单值行（合并标题/总分）与重复表头行。"""
    out = []
    seen = set()
    for i in range(start, min(start + 200, len(rows))):
        row = rows[i]
        if _skip_sample_row(row, header_row):
            continue
        if col >= len(row):
            continue
        v = row[col]
        if not v:
            continue
        if v in seen:
            continue
        seen.add(v)
        out.append(v)
        if len(out) >= limit:
            break
    return out


def _skip_sample_row(row: list[str], header_row: list[str] | None) -> bool:
    cells = [cell_text(c) for c in row if cell_text(c)]
    if not cells:
        return True  # 空行：跳过（不中断，兼容合并单元格产生的空行）
    if header_row and _header_overlap_ratio(row, header_row) >= 0.6:
        return True  # 多区块表格中的重复表头行
    joined = "".join(cells)
    if any(k in joined for k in FOOTER_KEYWORDS):
        return True  # 签字/合计等页脚行
    if len(cells) == 1:
        return True  # 合并单元格形成的单值行（区块标题/总分行）
    return False


def _header_overlap_ratio(row: list[str], header_row: list[str]) -> float:
    """计算数据行与表头行的重合度（表头非空单元格中相等的比例）。"""
    total = 0
    match = 0
    for i in range(min(len(row), len(header_row))):
        h = cell_text(header_row[i])
        if not h:
            continue
        total += 1
        if cell_text(row[i]) == h:
            match += 1
    return match / total if total else 0.0


def first_non_empty(*values: str) -> str:
    for v in values:
        if v is not None and str(v).strip():
            return str(v).strip()
    return ""


def title_from_filename(filename: str) -> str:
    base = Path(filename).stem
    return base.strip()


def main_guard() -> None:
    if sys.version_info < (3, 9):
        sys.exit("需要 Python 3.9+")
